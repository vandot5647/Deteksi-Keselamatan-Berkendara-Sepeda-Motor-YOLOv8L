"""
Aplikasi Deteksi Keselamatan Bermotor
-------------------------------------
Aplikasi ini menggunakan YOLO untuk mendeteksi keselamatan pengendara motor
melalui webcam atau video file. Aplikasi dapat mendeteksi:
- Penggunaan helm
- Pakaian tertutup/terbuka
- Penggunaan sepatu
"""

import sys
import os
import time
import cv2
import pygame
from datetime import datetime
from ultralytics import YOLO
from PyQt6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, 
    QPushButton, QLabel, QComboBox, QFileDialog, 
    QFrame, QHBoxLayout, QScrollArea, QSizePolicy, QMessageBox
)
from PyQt6.QtCore import Qt, QThread, pyqtSignal
from PyQt6.QtGui import QFont, QImage, QPixmap
import openpyxl
from openpyxl.styles import Font, Alignment
from typing import Dict
import numpy as np

# Konstanta
DETECTION_CLASSES = [
    'Helm', 'Pakai Sepatu', 'Pakaian Terbuka',
    'Pakaian Tertutup', 'Tidak Pakai Helm', 'Tidak Pakai Sepatu'
]

SAFETY_CLASSES = ['Pakaian Terbuka', 'Tidak Pakai Helm', 'Tidak Pakai Sepatu']
GOOD_CLASSES = {'Helm', 'Pakaian Tertutup', 'Pakai Sepatu'}

CONFIDENCE_THRESHOLD = 0.9
GOOD_CLASSES_THRESHOLD = 0.9
SCREENSHOT_INTERVAL = 0.3
VOICE_INTERVAL = 8.0

class DetectionThread(QThread):
    """Thread untuk menjalankan deteksi objek secara asynchronous"""
    
    update_frame = pyqtSignal(object)
    finished = pyqtSignal()
    error_occurred = pyqtSignal(str)  # Signal untuk error handling
    
    def __init__(self, model_path, source_type, video_path, screenshot_dir):
        super().__init__()
        self.model_path = model_path
        self.source_type = source_type
        self.video_path = video_path
        self.screenshot_dir = screenshot_dir
        self.running = True
        
        # Inisialisasi waktu deteksi terakhir
        self.last_detection_time: Dict[str, float] = {
            'Pakaian Terbuka': 0.0,
            'Tidak Pakai Helm': 0.0,
            'Tidak Pakai Sepatu': 0.0
        }
        
        # Inisialisasi counter untuk deteksi
        self.detection_count = {
            'Pakaian Terbuka': 0,
            'Tidak Pakai Helm': 0,
            'Tidak Pakai Sepatu': 0,
            'Helm': 0,
            'Pakaian Tertutup': 0,
            'Pakai Sepatu': 0,
        }
        
        # Setup audio
        self._setup_audio()
        
        # Inisialisasi log file
        self._setup_log_file()
        
    def _setup_audio(self):
        """Menyiapkan sistem audio dan file suara"""
        try:
            pygame.mixer.init()
            
            # Cari direktori Voice secara otomatis
            self.voice_path = self._find_voice_directory()
            
            if self.voice_path is not None:
                self.voice_files = {
                    'Pakaian Terbuka': os.path.join(self.voice_path, 'pakaian.mp3'),
                    'Tidak Pakai Helm': os.path.join(self.voice_path, 'helm.mp3'),
                    'Tidak Pakai Sepatu': os.path.join(self.voice_path, 'sepatu.mp3')
                }
                
                # Verifikasi file suara yang tersedia
                available_files = {}
                for class_name, file_path in self.voice_files.items():
                    if os.path.exists(file_path):
                        available_files[class_name] = file_path
                    else:
                        print(f"File suara tidak ditemukan: {file_path}")
                
                self.voice_files = available_files
            else:
                self.voice_files = {}
                print("Direktori Voice tidak ditemukan")
                
            self.last_voice_time: Dict[str, float] = {class_name: 0.0 for class_name in SAFETY_CLASSES}
        except Exception as e:
            # Jika audio gagal diinisialisasi, set flag untuk tidak menggunakan audio
            self.voice_files = {}
            self.last_voice_time = {}
            print(f"Error inisialisasi audio: {e}")

    def _find_voice_directory(self):
        """Mencari direktori Voice secara otomatis"""
        # Daftar kemungkinan lokasi direktori Voice
        possible_paths = [
            # Relatif terhadap direktori script saat ini
            os.path.join(os.path.dirname(os.path.abspath(__file__)), 'Voice'),
            # Relatif terhadap direktori kerja saat ini
            os.path.join(os.getcwd(), 'Voice'),
            # Cari di direktori parent
            os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), 'Voice'),
            # Cari di direktori yang sama dengan script
            os.path.join(os.path.dirname(os.path.abspath(__file__)), '..', 'Voice'),
            # Cari di folder _internal (untuk PyInstaller EXE)
            os.path.join(os.path.dirname(os.path.abspath(__file__)), '_internal', 'Voice'),
            # Cari di folder dist (untuk PyInstaller)
            os.path.join(os.path.dirname(os.path.abspath(__file__)), 'dist', 'Voice'),
        ]
        
        # Cari direktori yang berisi file MP3
        for path in possible_paths:
            if os.path.exists(path) and os.path.isdir(path):
                # Cek apakah ada file MP3 di direktori tersebut
                mp3_files = [f for f in os.listdir(path) if f.lower().endswith('.mp3')]
                if mp3_files:
                    print(f"Direktori Voice ditemukan: {path}")
                    print(f"File MP3 yang tersedia: {mp3_files}")
                    return path
        
        # Jika tidak ditemukan, cari secara rekursif di direktori script
        script_dir = os.path.dirname(os.path.abspath(__file__))
        for root, dirs, files in os.walk(script_dir):
            if 'Voice' in dirs:
                voice_dir = os.path.join(root, 'Voice')
                mp3_files = [f for f in os.listdir(voice_dir) if f.lower().endswith('.mp3')]
                if mp3_files:
                    print(f"Direktori Voice ditemukan (pencarian rekursif): {voice_dir}")
                    print(f"File MP3 yang tersedia: {mp3_files}")
                    return voice_dir
        
        return None

    def play_voice(self, class_name):
        """Memutar suara peringatan untuk kelas yang terdeteksi"""
        # Cek apakah audio tersedia
        if not self.voice_files or class_name not in self.voice_files:
            return
            
        current_time = time.time()
        if current_time - self.last_voice_time.get(class_name, 0.0) >= VOICE_INTERVAL:
            try:
                # Cek apakah file suara ada
                if os.path.exists(self.voice_files[class_name]):
                    pygame.mixer.music.load(self.voice_files[class_name])
                    pygame.mixer.music.play()
                    self.last_voice_time[class_name] = current_time
                    print(f"Memutar suara untuk: {class_name}")
                else:
                    print(f"File suara tidak ditemukan: {self.voice_files[class_name]}")
            except Exception as e:
                print(f"Error memutar suara untuk {class_name}: {e}")

    def _setup_log_file(self):
        """Menyiapkan file log Excel untuk menyimpan hasil deteksi"""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        self.log_file_path = os.path.join(self.screenshot_dir, f'detection_log_{timestamp}.xlsx')
        self.wb = openpyxl.Workbook()
        self.ws = self.wb.active
        if self.ws is not None:
            self.ws.title = "Log Deteksi"
            # Header
            headers = ["Waktu", "Jenis Deteksi", "Confidence", "Path Screenshot"]
            self.ws.append(headers)
            for col in range(1, len(headers)+1):
                self.ws.cell(row=1, column=col).font = Font(bold=True)
                self.ws.cell(row=1, column=col).alignment = Alignment(horizontal="center")
        self.wb.save(self.log_file_path)

    def _write_to_log(self, waktu, jenis, confidence, path):
        """Menulis baris ke file log Excel (hanya di memori, tidak langsung save ke file)"""
        try:
            if self.ws is not None:
                self.ws.append([waktu, jenis, confidence, path])
                # self.wb.save(self.log_file_path)  # Jangan save di sini!
        except Exception as e:
            pass

    def _process_detection(self, frame, result, current_time):
        boxes = result.boxes
        if len(boxes) == 0:
            return

        for box in boxes:
            xmin, ymin, xmax, ymax = box.xyxy[0].cpu().numpy()
            confidence = float(box.conf[0])
            class_idx = int(box.cls[0])
            class_name = DETECTION_CLASSES[class_idx]

            color = (0, 0, 255) if class_name in SAFETY_CLASSES else (255, 0, 0)
            cv2.rectangle(frame, (int(xmin), int(ymin)), (int(xmax), int(ymax)), color, 2)
            text = f'{class_name} {confidence:.2f}'
            cv2.putText(frame, text, (int(xmin), int(ymin - 10)), 
                       cv2.FONT_HERSHEY_SIMPLEX, 0.5, color, 2)

            # Log dan screenshot kelas tidak aman
            threshold = 0.85 if class_name == 'Tidak Pakai Helm' else CONFIDENCE_THRESHOLD
            if class_name in SAFETY_CLASSES and threshold <= confidence <= 1.0:
                self._handle_safety_detection(frame, class_name, confidence, current_time)
                self.play_voice(class_name)

            # Log kelas baik secara terpisah (log excel saja, tidak screenshot)
            if class_name in GOOD_CLASSES and confidence >= GOOD_CLASSES_THRESHOLD:
                if current_time - self.last_detection_time.get(class_name, 0.0) > SCREENSHOT_INTERVAL:
                    self.detection_count[class_name] += 1
                    waktu = datetime.now().strftime('%H:%M:%S')
                    self._write_to_log(waktu, class_name, f"{confidence:.2f}", "-")
                    self.last_detection_time[class_name] = current_time

        # Setelah loop, cek apakah SEMUA kelas baik terdeteksi bersamaan
        detected_classes = set(DETECTION_CLASSES[i] for i in result.boxes.cls.cpu().numpy().astype(int))
        if GOOD_CLASSES.issubset(detected_classes):
            # Cek confidence semua kelas baik
            all_good = True
            for class_name in GOOD_CLASSES:
                class_idx = DETECTION_CLASSES.index(class_name)
                confs = result.boxes.conf[result.boxes.cls.cpu().numpy().astype(int) == class_idx]
                if len(confs) == 0 or confs.max() < GOOD_CLASSES_THRESHOLD:
                    all_good = False
                    break
            if all_good and (current_time - self.last_detection_time.get('good_classes', 0.0) > SCREENSHOT_INTERVAL):
                # REMOVE: _save_good_classes_screenshot function
                pass # No screenshot for good classes anymore

    def _handle_safety_detection(self, frame, class_name, confidence, current_time):
        """Menangani deteksi kelas tidak aman"""
        if current_time - self.last_detection_time[class_name] > SCREENSHOT_INTERVAL:
            try:
                screenshot_path = os.path.join(
                    self.screenshot_dir, 'bad',
                    f'screenshot_{class_name}_{int(current_time)}.png'
                )
                cv2.imwrite(screenshot_path, frame.copy())
                self.last_detection_time[class_name] = current_time
                self.detection_count[class_name] += 1

                waktu = datetime.now().strftime('%H:%M:%S')
                self._write_to_log(waktu, class_name, f"{confidence:.2f}", screenshot_path)
            except Exception as e:
                # Jika gagal menyimpan screenshot, tetap log ke Excel
                waktu = datetime.now().strftime('%H:%M:%S')
                self._write_to_log(waktu, class_name, f"{confidence:.2f}", f"Error: {str(e)}")

    def run(self):
        """Menjalankan proses deteksi"""
        try:
            # Cek apakah file model ada
            if not os.path.exists(self.model_path):
                error_msg = "File model tidak ditemukan. Silakan pilih file model yang valid."
                self.error_occurred.emit(error_msg)
                self._write_to_log("Error", error_msg, "", "")
                return
            
            # Validasi ekstensi file model
            model_ext = os.path.splitext(self.model_path)[1].lower()
            if model_ext not in ['.pt', '.engine']:
                error_msg = f"Format file model '{model_ext}' tidak didukung. Silakan pilih model dengan format .pt (PyTorch) atau .engine."
                self.error_occurred.emit(error_msg)
                self._write_to_log("Error", error_msg, "", "")
                return
                
            # Coba load model
            try:
                model = YOLO(self.model_path)
                # Test model dengan dummy input untuk memastikan model dapat digunakan
                dummy_input = np.zeros((640, 640, 3), dtype=np.uint8)
                test_result = model(dummy_input, verbose=False)
            except Exception as model_error:
                error_msg = f"Gagal memuat model: {str(model_error)}. Silakan pilih model yang valid dan kompatibel."
                self.error_occurred.emit(error_msg)
                self._write_to_log("Error", error_msg, "", "")
                return
            
            # Buat direktori screenshot jika belum ada (hanya 'bad')
            os.makedirs(os.path.join(self.screenshot_dir, 'bad'), exist_ok=True)
                
            # Buka video capture
            cap = cv2.VideoCapture(0 if self.source_type == 'webcam' else self.video_path)
            if not cap.isOpened():
                error_msg = "Tidak dapat membuka video source. Pastikan webcam terhubung atau file video valid."
                self.error_occurred.emit(error_msg)
                self._write_to_log("Error", error_msg, "", "")
                return

            # Dapatkan resolusi asli video/webcam
            original_width = int(cap.get(cv2.CAP_PROP_FRAME_WIDTH))
            original_height = int(cap.get(cv2.CAP_PROP_FRAME_HEIGHT))
            fps_source = cap.get(cv2.CAP_PROP_FPS)
            
            print(f"Resolusi video asli: {original_width}x{original_height}")
            print(f"FPS video: {fps_source}")
            print("Model akan menggunakan resolusi asli video untuk deteksi yang lebih akurat")

            prev_time = time.time()
            fps = 0
            fps_update_time = time.time()
            frame_count = 0
            
            # Tambahkan timeout untuk deteksi model .engine
            frames_received = 0
            model_working = False

            while self.running:
                ret, frame = cap.read()
                if not ret:
                    if self.source_type == 'video':
                        self._write_to_log("Video selesai diputar", "", "", "")
                        break
                    else:
                        error_msg = "Tidak dapat membaca frame dari webcam. Pastikan webcam tidak digunakan aplikasi lain."
                        self.error_occurred.emit(error_msg)
                        break

                # Hitung FPS setiap 2 detik
                current_time = time.time()
                frame_count += 1
                frames_received += 1
                
                if current_time - fps_update_time >= 2.0:
                    fps = frame_count / (current_time - fps_update_time)
                    frame_count = 0
                    fps_update_time = current_time

                # Deteksi dengan resolusi asli video (PyTorch akan handle scaling otomatis)
                try:
                    results = model(frame, verbose=False)
                    for result in results:
                        self._process_detection(frame, result, current_time)
                    
                    # Jika berhasil melakukan deteksi, tandai model berfungsi
                    if not model_working:
                        model_working = True
                        
                except Exception as detection_error:
                    error_msg = f"Error saat deteksi: {str(detection_error)}. Model mungkin tidak kompatibel."
                    self.error_occurred.emit(error_msg)
                    break

                # Tampilkan informasi resolusi dan FPS di frame
                info_text = f"FPS: {int(fps)} | Video: {original_width}x{original_height}"
                cv2.putText(frame, info_text, (60, 30), cv2.FONT_HERSHEY_SIMPLEX, 0.7, (0,255,0), 2)
                self.update_frame.emit(frame)
                

            cap.release()
            if self.source_type == 'video':
                self.stop()
            self.finished.emit()  # Pastikan sinyal finished dipancarkan setelah selesai
        except Exception as e:
            error_msg = f"Error tidak terduga: {str(e)}"
            self.error_occurred.emit(error_msg)
            self._write_to_log("Error", error_msg, "", "")
            self.finished.emit()  # Juga emit jika error

    def stop(self):
        """Menghentikan proses deteksi"""
        if not self.running:
            return

        self.running = False
        pygame.mixer.quit()

        try:
            summary = self.wb.create_sheet("Kesimpulan")
            total_bad = sum(self.detection_count[class_name] for class_name in SAFETY_CLASSES)
            total_good = sum(self.detection_count[class_name] for class_name in GOOD_CLASSES)
            summary.append(["Waktu Selesai", datetime.now().strftime('%Y-%m-%d %H:%M:%S')])
            summary.append([])
            summary.append(["Total Deteksi Kelas Tidak Aman"])
            for class_name in SAFETY_CLASSES:
                summary.append([class_name, self.detection_count[class_name]])
            summary.append(["Total Kelas Tidak Aman", total_bad])
            summary.append([])
            summary.append(["Total Deteksi Kelas Baik"])
            for class_name in GOOD_CLASSES:
                summary.append([class_name, self.detection_count[class_name]])
            summary.append(["Total Kelas Baik", total_good])
            summary.append([])
            if total_bad + total_good > 0:
                compliance_rate = (total_good / (total_bad + total_good)) * 100
                summary.append(["Tingkat Kepatuhan (%)", f"{compliance_rate:.2f}"])
            if self.ws is not None:
                self._auto_fit_columns(self.ws)
            self._auto_fit_columns(summary)
            self.wb.save(self.log_file_path)
        except Exception as e:
            pass

    def _auto_fit_columns(self, ws):
        """Menyesuaikan lebar kolom sesuai isi terpanjang di setiap kolom"""
        if ws is None:
            return
        for column_cells in ws.columns:
            max_length = 0
            column = column_cells[0].column_letter  # Get the column name
            for cell in column_cells:
                try:
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
                except:
                    pass
            adjusted_width = max_length + 2  # Tambah sedikit padding
            ws.column_dimensions[column].width = adjusted_width

class MainWindow(QMainWindow):
    """Window utama aplikasi deteksi keselamatan sepeda motor"""
    
    def __init__(self):
        super().__init__()
        self._init_ui()
        self.detection_thread = None
        self.video_finished = False  # Tambahkan flag status video selesai
----
-
-
    --
    -
    -
